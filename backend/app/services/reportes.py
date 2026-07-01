"""
Servicio de reportes exportables — InvSystem Pro
Genera archivos Excel (.xlsx) y PDF (.pdf) con:
  - Reporte de inventario completo
  - Reporte de movimientos
  - Reporte de alertas (stock bajo / sin stock)
  - Reporte de resumen ejecutivo
"""
from __future__ import annotations

import io
from datetime import datetime, date
from typing import Any

# ── Excel ────────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ── PDF ──────────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, KeepTogether
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


# ── Paleta de colores ────────────────────────────────────────────
C_PRIMARY   = "1D9E75"
C_DARK      = "085041"
C_LIGHT     = "E1F5EE"
C_WARN      = "BA7517"
C_WARN_L    = "FAEEDA"
C_DANGER    = "993C1D"
C_DANGER_L  = "FAECE7"
C_GRAY      = "F6F6F4"
C_BORDER    = "E8E8E4"
C_TEXT      = "1A1A1A"
C_MUTED     = "6B6B6B"


# ════════════════════════════════════════════════════════════════
# EXCEL
# ════════════════════════════════════════════════════════════════
def _xl_header(ws, row: int, cols: list[str], bg: str = C_PRIMARY):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=True, color="FFFFFF", size=10)
    for c, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _xl_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _xl_autowidth(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def generar_excel(
    productos: list[dict],
    movimientos: list[dict],
    alertas: dict,
) -> bytes:
    wb = openpyxl.Workbook()

    # ── Hoja 1: Inventario ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Inventario"

    # Título
    ws1.merge_cells("A1:H1")
    t = ws1["A1"]
    t.value = f"InvSystem Pro — Reporte de Inventario  |  {date.today().strftime('%d/%m/%Y')}"
    t.font = Font(bold=True, size=13, color=C_DARK)
    t.fill = PatternFill("solid", fgColor=C_LIGHT)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    cols_inv = ["SKU", "Producto", "Categoría", "Stock actual", "Stock mínimo", "Precio unit.", "Valor total", "Estado"]
    _xl_header(ws1, 2, cols_inv)

    for i, p in enumerate(productos, 3):
        valor = p["stock_actual"] * p["precio_unitario"]
        if p["stock_actual"] == 0:
            estado, color = "Sin stock", C_DANGER_L
        elif p["stock_actual"] < p["stock_minimo"]:
            estado, color = "Stock bajo", C_WARN_L
        else:
            estado, color = "Normal", C_LIGHT

        row_data = [
            p["sku"], p["nombre"], p["categoria"],
            p["stock_actual"], p["stock_minimo"],
            f"${p['precio_unitario']:,.2f}",
            f"${valor:,.2f}", estado,
        ]
        fill = PatternFill("solid", fgColor=color if i % 2 == 0 else "FFFFFF")
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=i, column=c, value=val)
            cell.border = _xl_border()
            cell.alignment = Alignment(vertical="center")
            if estado != "Normal" and c == 8:
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(
                    bold=True,
                    color=C_DANGER if estado == "Sin stock" else C_WARN
                )

    # Totales
    last = len(productos) + 2
    ws1.cell(row=last + 1, column=1, value="TOTALES").font = Font(bold=True)
    total_val = sum(p["stock_actual"] * p["precio_unitario"] for p in productos)
    ws1.cell(row=last + 1, column=7, value=f"${total_val:,.2f}").font = Font(bold=True, color=C_DARK)

    _xl_autowidth(ws1)
    ws1.freeze_panes = "A3"

    # ── Hoja 2: Movimientos ─────────────────────────────────────
    ws2 = wb.create_sheet("Movimientos")
    ws2.merge_cells("A1:F1")
    t2 = ws2["A1"]
    t2.value = f"InvSystem Pro — Movimientos  |  {date.today().strftime('%d/%m/%Y')}"
    t2.font = Font(bold=True, size=13, color=C_DARK)
    t2.fill = PatternFill("solid", fgColor=C_LIGHT)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    cols_mov = ["Fecha", "Producto", "Tipo", "Cantidad", "Stock resultante", "Motivo"]
    _xl_header(ws2, 2, cols_mov)

    for i, m in enumerate(movimientos, 3):
        tipo_color = {"entrada": C_LIGHT, "salida": C_DANGER_L, "ajuste": C_WARN_L}.get(m.get("tipo", ""), "FFFFFF")
        row_data = [
            m.get("fecha", "")[:10] if m.get("fecha") else "",
            m.get("producto_nombre", m.get("producto_id", "")),
            m.get("tipo", "").upper(),
            m.get("cantidad", 0),
            m.get("stock_resultante", 0),
            m.get("motivo", ""),
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=i, column=c, value=val)
            cell.border = _xl_border()
            if c == 3:
                cell.fill = PatternFill("solid", fgColor=tipo_color)
                cell.font = Font(bold=True)

    _xl_autowidth(ws2)
    ws2.freeze_panes = "A3"

    # ── Hoja 3: Resumen ─────────────────────────────────────────
    ws3 = wb.create_sheet("Resumen ejecutivo")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 18

    ws3.merge_cells("A1:B1")
    t3 = ws3["A1"]
    t3.value = "Resumen ejecutivo"
    t3.font = Font(bold=True, size=14, color=C_DARK)
    t3.fill = PatternFill("solid", fgColor=C_LIGHT)
    t3.alignment = Alignment(horizontal="center")
    ws3.row_dimensions[1].height = 30

    kpis = [
        ("Total de productos", len(productos)),
        ("Valor total del inventario", f"${sum(p['stock_actual']*p['precio_unitario'] for p in productos):,.2f}"),
        ("Productos sin stock", alertas.get("resumen", {}).get("sin_stock", 0)),
        ("Productos con stock bajo", alertas.get("resumen", {}).get("stock_bajo", 0)),
        ("Productos en estado normal", alertas.get("resumen", {}).get("normal", 0)),
        ("Fecha del reporte", date.today().strftime("%d/%m/%Y")),
    ]

    for i, (label, val) in enumerate(kpis, 3):
        ws3.cell(row=i, column=1, value=label).font = Font(color=C_MUTED)
        cell = ws3.cell(row=i, column=2, value=val)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="right")
        for c in [1, 2]:
            ws3.cell(row=i, column=c).border = _xl_border()
            if i % 2 == 0:
                ws3.cell(row=i, column=c).fill = PatternFill("solid", fgColor=C_GRAY)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════
# PDF
# ════════════════════════════════════════════════════════════════
def generar_pdf(
    productos: list[dict],
    movimientos: list[dict],
    alertas: dict,
    tipo: str = "inventario",
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4) if tipo == "inventario" else A4,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    style_title = ParagraphStyle(
        "titulo", fontSize=16, fontName="Helvetica-Bold",
        textColor=colors.HexColor(f"#{C_DARK}"), spaceAfter=4,
    )
    style_sub = ParagraphStyle(
        "sub", fontSize=10, fontName="Helvetica",
        textColor=colors.HexColor(f"#{C_MUTED}"), spaceAfter=12,
    )
    style_section = ParagraphStyle(
        "seccion", fontSize=12, fontName="Helvetica-Bold",
        textColor=colors.HexColor(f"#{C_DARK}"),
        spaceBefore=12, spaceAfter=6,
    )

    story = []

    # ── Encabezado ───────────────────────────────────────────────
    story.append(Paragraph("InvSystem Pro", style_title))
    story.append(Paragraph(
        f"Reporte de {tipo.title()}  —  Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        style_sub
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor(f"#{C_PRIMARY}")))
    story.append(Spacer(1, 0.4 * cm))

    # ── KPIs ─────────────────────────────────────────────────────
    total_val = sum(p["stock_actual"] * p["precio_unitario"] for p in productos)
    res = alertas.get("resumen", {})
    kpi_data = [
        ["Total productos", "Valor inventario", "Sin stock", "Stock bajo", "Normal"],
        [
            str(len(productos)),
            f"${total_val:,.0f}",
            str(res.get("sin_stock", 0)),
            str(res.get("stock_bajo", 0)),
            str(res.get("normal", 0)),
        ],
    ]
    kpi_table = Table(kpi_data, colWidths=[4 * cm] * 5)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{C_PRIMARY}")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 9),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor(f"#{C_LIGHT}")),
        ("FONTNAME",   (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 1), (-1, 1), 13),
        ("TEXTCOLOR",  (2, 1), (2, 1), colors.HexColor(f"#{C_DANGER}")),
        ("TEXTCOLOR",  (3, 1), (3, 1), colors.HexColor(f"#{C_WARN}")),
        ("TEXTCOLOR",  (4, 1), (4, 1), colors.HexColor(f"#{C_DARK}")),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [None, None]),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor(f"#{C_BORDER}")),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.5 * cm))

    # ── Tabla de inventario ──────────────────────────────────────
    if tipo in ("inventario", "completo"):
        story.append(Paragraph("Detalle de inventario", style_section))

        headers = ["SKU", "Producto", "Categoría", "Stock", "Mín.", "Precio", "Valor", "Estado"]
        rows = [headers]
        for p in productos:
            val = p["stock_actual"] * p["precio_unitario"]
            if p["stock_actual"] == 0:
                estado = "SIN STOCK"
            elif p["stock_actual"] < p["stock_minimo"]:
                estado = "BAJO"
            else:
                estado = "OK"
            rows.append([
                p["sku"], p["nombre"][:28], p["categoria"],
                str(p["stock_actual"]), str(p["stock_minimo"]),
                f"${p['precio_unitario']:,.2f}",
                f"${val:,.2f}", estado,
            ])

        col_w = [2.2*cm, 6*cm, 3*cm, 1.8*cm, 1.8*cm, 2.5*cm, 2.8*cm, 2*cm]
        tbl = Table(rows, colWidths=col_w, repeatRows=1)

        ts = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{C_PRIMARY}")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("ALIGN",      (3, 0), (7, -1), "CENTER"),
            ("ALIGN",      (0, 0), (2, -1), "LEFT"),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("GRID",       (0, 0), (-1, -1), 0.3, colors.HexColor(f"#{C_BORDER}")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
                colors.white, colors.HexColor(f"#{C_GRAY}")
            ]),
        ])

        # Colorear filas por estado
        for i, p in enumerate(productos, 1):
            if p["stock_actual"] == 0:
                ts.add("BACKGROUND", (7, i), (7, i), colors.HexColor(f"#{C_DANGER_L}"))
                ts.add("TEXTCOLOR",  (7, i), (7, i), colors.HexColor(f"#{C_DANGER}"))
                ts.add("FONTNAME",   (7, i), (7, i), "Helvetica-Bold")
            elif p["stock_actual"] < p["stock_minimo"]:
                ts.add("BACKGROUND", (7, i), (7, i), colors.HexColor(f"#{C_WARN_L}"))
                ts.add("TEXTCOLOR",  (7, i), (7, i), colors.HexColor(f"#{C_WARN}"))
                ts.add("FONTNAME",   (7, i), (7, i), "Helvetica-Bold")

        tbl.setStyle(ts)
        story.append(tbl)

    # ── Alertas ──────────────────────────────────────────────────
    if tipo in ("alertas", "completo"):
        sin_stock = alertas.get("sin_stock", [])
        stock_bajo = alertas.get("stock_bajo", [])

        if sin_stock:
            story.append(Spacer(1, 0.4 * cm))
            story.append(Paragraph("Productos sin stock", style_section))
            data = [["SKU", "Producto", "Categoría", "Mínimo requerido"]]
            for p in sin_stock:
                data.append([p["sku"], p["nombre"], p["categoria"], str(p["stock_minimo"])])
            t = Table(data, colWidths=[3*cm, 8*cm, 4*cm, 4*cm])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{C_DANGER}")),
                ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
                ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",   (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor(f"#{C_DANGER_L}")),
                ("GRID",       (0, 0), (-1, -1), 0.3, colors.HexColor(f"#{C_BORDER}")),
                ("TOPPADDING",    (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(t)

        if stock_bajo:
            story.append(Spacer(1, 0.3 * cm))
            story.append(Paragraph("Productos con stock bajo", style_section))
            data2 = [["SKU", "Producto", "Stock actual", "Mínimo", "Diferencia"]]
            for p in stock_bajo:
                data2.append([
                    p["sku"], p["nombre"],
                    str(p["stock_actual"]), str(p["stock_minimo"]),
                    str(p["stock_minimo"] - p["stock_actual"]),
                ])
            t2 = Table(data2, colWidths=[3*cm, 7*cm, 3*cm, 3*cm, 3*cm])
            t2.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{C_WARN}")),
                ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
                ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",   (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor(f"#{C_WARN_L}")),
                ("GRID",       (0, 0), (-1, -1), 0.3, colors.HexColor(f"#{C_BORDER}")),
                ("TOPPADDING",    (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(t2)

    doc.build(story)
    return buf.getvalue()
